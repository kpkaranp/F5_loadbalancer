#!/usr/bin/env python
"""
Script to generate a report of F5 virtual servers with pool and node information
by fetching data directly from F5 API endpoints.

This script fetches data from the following F5 endpoints:
- /mgmt/tm/ltm/virtual
- /mgmt/tm/ltm/virtual/stats
- /mgmt/tm/ltm/pool
- /mgmt/tm/ltm/pool/stats
- /mgmt/tm/ltm/pool/members
- /mgmt/tm/ltm/node
- /mgmt/tm/ltm/node/stats

Generates Excel report with detailed information and summary statistics.
"""

import requests
import json
import os
from datetime import datetime
import re
from collections import defaultdict, Counter
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
import argparse
import logging
import sys

# Suppress insecure HTTPS warnings
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('f5_status_report.log')
    ]
)
logger = logging.getLogger(__name__)

class F5Config:
    """Client for interacting with F5 API."""
    
    def __init__(self, host, username, password, verify_ssl=False):
        """Initialize F5 client with connection details and generate report."""
        # Ensure host has https:// prefix
        if not host.startswith('http'):
            self.F5_HOST = f"https://{host}"
        else:
            self.F5_HOST = host
            
        self.USERNAME = username
        self.PASSWORD = password
        
        # Create session
        self.session = requests.Session()
        self.session.auth = HTTPBasicAuth(self.USERNAME, self.PASSWORD)
        self.session.verify = verify_ssl
        self.session.headers.update({'Content-Type': 'application/json'})

        # Automatically generate report upon initialization
        try:
            logger.info("Fetching virtual server data...")
            vs_data, summary_counts = process_virtual_servers(self)
            
            logger.info("Fetching pool data...")
            pool_data = process_pools(self, summary_counts)
            
            logger.info("Fetching node data...")
            node_data = process_nodes(self, summary_counts)
            
            # Generate report
            logger.info("Generating report...")
            report_data = generate_report(vs_data, pool_data, node_data)
            
            # Generate output filename prefix
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            f5_hostname = self.F5_HOST.replace("https://", "").replace("http://", "")
            output_prefix = f"{f5_hostname}_{timestamp}_f5_report"
            
            # Generate Excel report
            excel_filename = generate_excel_report(report_data, summary_counts, output_prefix)
            logger.info(f"Report generated in: {os.path.abspath(os.path.dirname(excel_filename))}")
            
            # Print summary
            print("\nSummary of F5 Components:")
            print("-" * 50)
            print("Virtual Servers:")
            for status, count in summary_counts['virtual'].items():
                print(f"  {status}: {count}")
            
            print("\nPools:")
            for status, count in summary_counts['pool'].items():
                print(f"  {status}: {count}")
            
            print("\nNodes:")
            for status, count in summary_counts['node'].items():
                print(f"  {status}: {count}")
            
            print(f"\nTotal virtual servers: {len(report_data)}")
            #print(f"Total pools with members: {sum(1 for item in report_data if item.get('active_members', 0) > 0)}")
            def safe_int(val):
                try:
                    return int(val)
                except (ValueError, TypeError):
                    return 0
            
            print(f"Total pools with members: {sum(1 for item in report_data if safe_int(item.get('active_members', 0)) > 0)}")
            print(f"Total pools: {len(pool_data)}")
            print(f"Total nodes: {len(node_data)}")
            
        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            raise

    def get_json(self, endpoint):
        """Get data from F5 API endpoint."""
        url = f"{self.F5_HOST}{endpoint}"
        try:
            resp = self.session.get(url)
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            logger.error(f"Failed to get data from {endpoint}: {str(e)}")
            return None

def extract_name_from_path(path):
    """Extract the name from a path, handling different formats."""
    if '~' in path:
        parts = path.split('~')
        if len(parts) > 1:
            return parts[-1]
    
    if '/' in path:
        parts = path.split('/')
        if len(parts) > 1:
            return parts[-1]
    
    return path

def parse_destination(destination):
    """Parse destination field to extract IP and port."""
    ip = ""
    port = ""
    
    if ':' in destination:
        parts = destination.rsplit(':', 1)
        if len(parts) == 2:
            ip_part = parts[0]
            port = parts[1]
            
            if '/' in ip_part:
                ip = ip_part.split('/')[-1]
            else:
                ip = ip_part
    else:
        if '/' in destination:
            ip = destination.split('/')[-1]
    
    # Remove route domain if present
    if '%' in ip:
        ip = ip.split('%')[0]
    
    return ip, port

def process_virtual_servers(f5_config):
    """Fetch and process virtual server information using robust data-driven mapping."""
    try:
        virtuals = f5_config.get_json('/mgmt/tm/ltm/virtual')
        vstats = f5_config.get_json('/mgmt/tm/ltm/virtual/stats')
        if not virtuals or not vstats:
            logger.error("Failed to get virtual server data or stats")
            return [], {"virtual": Counter(), "pool": Counter(), "node": Counter()}
        
        # Build a map from fullPath to stats using tmName.description
        stats_map = {}
        for entry in vstats.get('entries', {}).values():
            nested = entry.get('nestedStats', {}).get('entries', {})
            tm_name = nested.get('tmName', {}).get('description', '')
            if tm_name:
                stats_map[tm_name] = nested
        
        vs_data = []
        summary_counts = {
            "virtual": Counter(),
            "pool": Counter(),
            "node": Counter()
        }
        for virtual in virtuals.get('items', []):
            try:
                name = virtual.get('name', '')
                fullPath = virtual.get('fullPath', '')
                partition = virtual.get('partition', '')
                stats = stats_map.get(fullPath, {})
                if not stats:
                    logger.warning(f"Could not get stats for virtual server: {name} ({fullPath})")
                    continue
                vs_info = {
                    'name': name,
                    'fullPath': fullPath,
                    'partition': partition,
                    'description': virtual.get('description', ''),
                    'destination_ip': '',
                    'destination_port': '',
                    'pool': virtual.get('pool', ''),
                    'availabilityState': stats.get('status.availabilityState', {}).get('description', 'N/A'),
                    'enabledState': stats.get('status.enabledState', {}).get('description', 'N/A'),
                    'statusReason': stats.get('status.statusReason', {}).get('description', 'N/A')
                }
                summary_counts["virtual"][vs_info['availabilityState']] += 1
                destination = virtual.get('destination', '')
                if destination:
                    ip, port = parse_destination(destination)
                    vs_info['destination_ip'] = ip
                    vs_info['destination_port'] = port
                vs_data.append(vs_info)
            except Exception as e:
                logger.error(f"Error processing virtual server {name}: {str(e)}")
                continue
        return vs_data, summary_counts
    except Exception as e:
        logger.error(f"Error in process_virtual_servers: {str(e)}")
        return [], {"virtual": Counter(), "pool": Counter(), "node": Counter()}

def process_pools(f5_config, summary_counts):
    """Fetch and process pool information, always fetching members live from membersReference.link, and mapping status by fullPath/tmName.description only."""
    try:
        pools = f5_config.get_json('/mgmt/tm/ltm/pool')
        pstats = f5_config.get_json('/mgmt/tm/ltm/pool/stats')
        if not pools or not pstats:
            logger.error("Failed to get pool data or stats")
            return {}
        # Build a map from fullPath to stats using tmName.description
        stats_map = {}
        for entry in pstats.get('entries', {}).values():
            nested = entry.get('nestedStats', {}).get('entries', {})
            tm_name = nested.get('tmName', {}).get('description', '')
            if tm_name:
                stats_map[tm_name] = nested
        pool_data = {}
        for pool in pools.get('items', []):
            try:
                fullPath = pool.get('fullPath', '')
                name = pool.get('name', '')
                partition = pool.get('partition', '')
                stats = stats_map.get(fullPath, {})
                if not stats:
                    logger.warning(f"Could not get stats for pool: {name} ({fullPath})")
                    continue
                availability_state = stats.get('status.availabilityState', {}).get('description', 'N/A')
                summary_counts["pool"][availability_state] += 1
                pool_data[fullPath] = {
                    'name': name,
                    'partition': partition,
                    'monitor': pool.get('monitor', ''),
                    'availabilityState': availability_state,
                    'enabledState': stats.get('status.enabledState', {}).get('description', 'N/A'),
                    'statusReason': stats.get('status.statusReason', {}).get('description', 'N/A'),
                    'activeMemberCount': stats.get('activeMemberCnt', {}).get('value', 0),
                    'totalMemberCount': stats.get('memberCnt', {}).get('value', 0),
                    'members': []
                }
                # Always fetch members live from membersReference.link
                members_ref = pool.get('membersReference', {}).get('link')
                if members_ref:
                    members_ref = members_ref.split('?')[0]
                    if members_ref.startswith('https://localhost'):
                        members_ref = members_ref.replace('https://localhost', '')
                    members_data = f5_config.get_json(members_ref)
                    if members_data and 'items' in members_data:
                        for member in members_data['items']:
                            pool_data[fullPath]['members'].append({
                                'name': member.get('name', ''),
                                'address': member.get('address', ''),
                                'port': member.get('port', ''),
                                'state': member.get('state', ''),
                                'session': member.get('session', '')
                            })
            except Exception as e:
                logger.error(f"Error processing pool {name}: {str(e)}")
                continue
        return pool_data
    except Exception as e:
        logger.error(f"Error in process_pools: {str(e)}")
        return {}

def process_nodes(f5_config, summary_counts):
    """Fetch and process node information using robust data-driven mapping."""
    try:
        nodes = f5_config.get_json('/mgmt/tm/ltm/node')
        nstats = f5_config.get_json('/mgmt/tm/ltm/node/stats')
        if not nodes or not nstats:
            logger.error("Failed to get node data or stats")
            return {}
        # Build a map from fullPath to stats using tmName.description
        stats_map = {}
        for entry in nstats.get('entries', {}).values():
            nested = entry.get('nestedStats', {}).get('entries', {})
            tm_name = nested.get('tmName', {}).get('description', '')
            if tm_name:
                stats_map[tm_name] = nested
        node_data = {}
        for node in nodes.get('items', []):
            try:
                name = node.get('name', '')
                address = node.get('address', '')
                partition = node.get('partition', '')
                fullPath = node.get('fullPath', '')
                stats = stats_map.get(fullPath, {})
                if not stats:
                    logger.warning(f"Could not get stats for node: {name} ({fullPath})")
                    continue
                availability_state = stats.get('status.availabilityState', {}).get('description', 'N/A')
                summary_counts["node"][availability_state] += 1
                node_data[address] = {
                    'name': name,
                    'fullPath': fullPath,
                    'partition': partition,
                    'availabilityState': availability_state,
                    'enabledState': stats.get('status.enabledState', {}).get('description', 'N/A'),
                    'statusReason': stats.get('status.statusReason', {}).get('description', 'N/A')
                }
            except Exception as e:
                logger.error(f"Error processing node {name}: {str(e)}")
                continue
        return node_data
    except Exception as e:
        logger.error(f"Error in process_nodes: {str(e)}")
        return {}

def generate_excel_report(report_data, summary_counts, output_prefix, pool_data):
    """Generate Excel report with summary and details."""
    try:
        # Create initial workbook
        wb = Workbook()
        
        # Create List sheet first (it will be second after we create Summary)
        list_sheet = wb.active
        list_sheet.title = "List"
        
        # Create Summary sheet and make it active
        summary_sheet = wb.create_sheet("Summary", 0)  # 0 index makes it the first sheet
        
        # Define color fills
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        
        # Add F5 hostname as heading
        f5_hostname = output_prefix.split('_')[0]  # Extract hostname from output_prefix
        summary_sheet['A1'] = f5_hostname
        summary_sheet.merge_cells('A1:F1')  # Merge cells from A1 to F1
        
        # Format the heading
        heading_cell = summary_sheet['A1']
        heading_cell.fill = green_fill
        heading_cell.font = Font(color="FFFFFF", bold=True, size=14)  # White color, bold, size 14
        heading_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Create summary data for DataFrame
        summary_data = []
        for k, counts in summary_counts.items():
            summary_data.append({
                'Type': k.capitalize(),
                'Available': counts.get("available", 0),
                'Offline': counts.get("offline", 0),
                'Online': counts.get("online", 0),
                'Unknown': counts.get("unknown", 0),
                'Total': sum(counts.values())
            })
        
        # Create DataFrame
        summary_df = pd.DataFrame(summary_data)
        
        # Save initial workbook
        excel_filename = f"{output_prefix}.xlsx"
        wb.save(excel_filename)
        
        # Write DataFrame to Excel
        with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a') as writer:
            writer.book = load_workbook(excel_filename)
            summary_df.to_excel(writer, sheet_name='Summary', startrow=1, startcol=0, index=False)
            writer.save()
        
        # Apply colors to the summary sheet
        book = load_workbook(excel_filename)
        ws = book['Summary']
        
        # Apply colors to each row in the summary
        for row in range(3, len(summary_data) + 3):  # Start from row 3 (after heading and headers)
            for col in range(1, 7):  # Columns A through F
                cell = ws.cell(row=row, column=col)
                if col == 1:  # Type column
                    cell.fill = blue_fill
                elif col == 2:  # Available column
                    cell.fill = green_fill
                elif col == 3:  # Offline column
                    cell.fill = red_fill
                elif col == 4:  # Online column
                    cell.fill = green_fill
                elif col == 5:  # Unknown column
                    cell.fill = yellow_fill
                elif col == 6:  # Total column
                    cell.fill = blue_fill
        
        # Add list headers with hierarchical structure
        headers = [
            # Virtual Server Information
            "Virtual Server", "VS Description", "VS Destination", "VS Service Port",
            "VS Status", "VS Status Reason",
            # Pool Information
            "Pool Name", "Pool Status", "Pool Status Reason",
            "Pool Active Members", "Pool Total Members",
            # Member Information
            "Member Name", "Member Address", "Member Port",
            "Member State", "Member Session",
            # Node Information
            "Node IP", "Node Status", "Node Status Reason"
        ]
        list_sheet = book['List']
        list_sheet.append(headers)
        
        # Add list data with hierarchical structure
        for data in report_data:
            pool_info = data.get('pool', '')
            node_names = data.get('node_names', '').split('; ')
            node_addresses = data.get('node_addresses', '').split('; ')
            node_availability_states = data.get('node_availability_states', '').split('; ')
            node_status_reasons = data.get('node_status_reasons', '').split('; ')
            
            # Get pool members from the pool data
            pool_members = []
            if pool_info in pool_data:
                pool_members = pool_data[pool_info].get('members', [])
            
            # If there are no nodes or pool members, add one row with blank values
            if node_names == [''] and not pool_members:
                row = [
                    # Virtual Server Information
                    data['name'],
                    data['description'],
                    f"{data['destination_ip']}:{data['destination_port']}" if data['destination_ip'] else '',
                    data['destination_port'],
                    data['vs_availabilityState'],
                    data['vs_statusReason'],
                    # Pool Information
                    data['pool_name'],
                    data['pool_availabilityState'],
                    data['pool_statusReason'],
                    data.get('active_members', ''),
                    data.get('total_members', ''),
                    # Member Information
                    '',  # Member Name
                    '',  # Member Address
                    '',  # Member Port
                    '',  # Member State
                    '',  # Member Session
                    # Node Information
                    '',  # Node IP
                    '',  # Node Status
                    ''   # Node Status Reason
                ]
                list_sheet.append(row)
            else:
                # Add a row for each pool member
                for member in pool_members:
                    row = [
                        # Virtual Server Information
                        data['name'],
                        data['description'],
                        f"{data['destination_ip']}:{data['destination_port']}" if data['destination_ip'] else '',
                        data['destination_port'],
                        data['vs_availabilityState'],
                        data['vs_statusReason'],
                        # Pool Information
                        data['pool_name'],
                        data['pool_availabilityState'],
                        data['pool_statusReason'],
                        data.get('active_members', ''),
                        data.get('total_members', ''),
                        # Member Information
                        member.get('name', ''),
                        member.get('address', ''),
                        member.get('port', ''),
                        member.get('state', ''),
                        member.get('session', ''),
                        # Node Information
                        member.get('address', ''),  # Node IP is same as member address
                        node_availability_states[0] if node_availability_states else '',
                        node_status_reasons[0] if node_status_reasons else ''
                    ]
                    list_sheet.append(row)
        
        # Save the final workbook
        book.save(excel_filename)
        logger.info(f"Excel report generated: {excel_filename}")
        return excel_filename
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise

def generate_report(vs_data, pool_data, node_data):
    """Generate combined report data."""
    report_data = []
    
    for vs in vs_data:
        pool_path = vs.get('pool', '')
        pool_info = pool_data.get(pool_path, {})
        
        # Get node information for pool members
        node_names = []
        node_addresses = []
        node_availability_states = []
        node_enabled_states = []
        node_status_reasons = []
        member_states = []
        member_sessions = []
        
        for member in pool_info.get('members', []):
            member_address = member.get('address', '')
            node_info = node_data.get(member_address, {})
            
            if node_info:
                node_names.append(node_info['name'])
                node_addresses.append(member_address)
                node_availability_states.append(node_info['availabilityState'])
                node_enabled_states.append(node_info['enabledState'])
                node_status_reasons.append(node_info['statusReason'])
                member_states.append(member.get('state', ''))
                member_sessions.append(member.get('session', ''))
        
        report_entry = {
            # Virtual server info
            'name': vs['name'],
            'fullPath': vs['fullPath'],
            'description': vs['description'],
            'destination_ip': vs['destination_ip'],
            'destination_port': vs['destination_port'],
            
            # Virtual server status
            'vs_availabilityState': vs['availabilityState'],
            'vs_enabledState': vs['enabledState'],
            'vs_statusReason': vs['statusReason'],
            
            # Pool info
            'pool': pool_path,
            'pool_name': pool_info.get('name', ''),
            'pool_monitor': pool_info.get('monitor', ''),
            'active_members': pool_info.get('activeMemberCount', ''),
            'total_members': pool_info.get('totalMemberCount', ''),
            
            # Pool status
            'pool_availabilityState': pool_info.get('availabilityState', ''),
            'pool_enabledState': pool_info.get('enabledState', ''),
            'pool_statusReason': pool_info.get('statusReason', ''),
            
            # Node information
            'node_names': '; '.join(node_names) if node_names else '',
            'node_addresses': '; '.join(node_addresses) if node_addresses else '',
            'node_availability_states': '; '.join(node_availability_states) if node_availability_states else '',
            'node_enabled_states': '; '.join(node_enabled_states) if node_enabled_states else '',
            'node_status_reasons': '; '.join(node_status_reasons) if node_status_reasons else '',
            
            # Member information
            'member_states': '; '.join(member_states) if member_states else '',
            'member_sessions': '; '.join(member_sessions) if member_sessions else ''
        }
        
        report_data.append(report_entry)
    
    return report_data

def main():
    parser = argparse.ArgumentParser(description='Generate F5 status report')
    parser.add_argument('--host', required=True, help='F5 hostname or IP address')
    parser.add_argument('--username', required=True, help='F5 username')
    parser.add_argument('--password', required=True, help='F5 password')
    parser.add_argument('--verify-ssl', action='store_true', help='Verify SSL certificate')
    
    args = parser.parse_args()
    
    try:
        # Initialize F5 config
        f5_config = F5Config(args.host, args.username, args.password, args.verify_ssl)
        
        # Fetch and process data
        logger.info("Fetching virtual server data...")
        vs_data, summary_counts = process_virtual_servers(f5_config)
        
        logger.info("Fetching pool data...")
        pool_data = process_pools(f5_config, summary_counts)
        
        logger.info("Fetching node data...")
        node_data = process_nodes(f5_config, summary_counts)
        
        # Generate report
        logger.info("Generating report...")
        report_data = generate_report(vs_data, pool_data, node_data)
        
        # Generate output filename prefix
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        f5_hostname = args.host.replace("https://", "").replace("http://", "")
        output_prefix = f"{f5_hostname}_{timestamp}_f5_report"
        
        # Generate Excel report
        excel_filename = generate_excel_report(report_data, summary_counts, output_prefix, pool_data)
        logger.info(f"Report generated in: {os.path.abspath(os.path.dirname(excel_filename))}")
        
        # Print summary
        print("\nSummary of F5 Components:")
        print("-" * 50)
        print("Virtual Servers:")
        for status, count in summary_counts['virtual'].items():
            print(f"  {status}: {count}")
        
        print("\nPools:")
        for status, count in summary_counts['pool'].items():
            print(f"  {status}: {count}")
        
        print("\nNodes:")
        for status, count in summary_counts['node'].items():
            print(f"  {status}: {count}")
        
        print(f"\nTotal virtual servers: {len(report_data)}")
        print(f"Total pools with members: {sum(1 for item in report_data if item.get('active_members', 0) > 0)}")
        print(f"Total pools: {len(pool_data)}")
        print(f"Total nodes: {len(node_data)}")
        
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        sys.exit(1)

if __name__ == "__main__":
    main() 
