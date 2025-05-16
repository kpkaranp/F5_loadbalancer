import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
import requests
from urllib3.exceptions import InsecureRequestWarning

def add_status_summary_to_excel(
    excel_path, summary_df, sheet_name="Status Summary", start_row=10, start_col=2, device_address=None
):
    """
    Adds a summary DataFrame to a new sheet in the Excel file, starting at (start_row, start_col),
    and colors status columns. Does not affect other sheets.
    """
    book = load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    writer.book = book
    summary_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row-1, startcol=start_col-1, index=False)
    writer.save()
    writer.close()
    ws = book[sheet_name]
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    nrows, ncols = summary_df.shape
    # Add device address header above the summary table
    if device_address:
        header_cell = ws.cell(row=start_row-1, column=start_col)
        header_cell.value = f"F5 Device: {device_address}"
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Merge header across all columns of the summary table
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=start_col+ncols-1)
    for row in ws.iter_rows(min_row=start_row+1, max_row=start_row+nrows, min_col=start_col, max_col=start_col+ncols-1):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if "available" in ws.cell(row=start_row, column=cell.column).value.lower():
                    cell.fill = green_fill
                elif "unavailable" in ws.cell(row=start_row, column=cell.column).value.lower():
                    cell.fill = yellow_fill
                elif "offline" in ws.cell(row=start_row, column=cell.column).value.lower():
                    cell.fill = red_fill
                elif "unknown" in ws.cell(row=start_row, column=cell.column).value.lower():
                    cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    book.save(excel_path)
    book.close()

def fetch_and_write_f5_summary_excel_with_token(address, token, excel_path):
    """
    Fetches F5 summary stats using an existing token and writes a summary sheet to the given Excel file.
    Usage:
        from f5_status_summary_sheet import fetch_and_write_f5_summary_excel_with_token
        fetch_and_write_f5_summary_excel_with_token(address, token, excel_path)
    """
    def get_stats(url, token):
        headers = {'X-F5-Auth-Token': token, 'Content-Type': 'application/json'}
        r = requests.get(url, headers=headers, verify=False)
        r.raise_for_status()
        return r.json()

    def parse_stats_entries(stats_json):
        entries = stats_json.get('entries', {})
        total = len(entries)
        available = unavailable = offline = unknown = 0
        available_disabled = offline_disabled = unknown_disabled = 0
        for entry in entries.values():
            nested = entry.get('nestedStats', {}).get('entries', {})
            status = nested.get('status.availabilityState', {}).get('description', '').lower()
            enabled = nested.get('status.enabledState', {}).get('description', '').lower()
            disabled = enabled == 'disabled'
            if status == 'available':
                available += 1
                if disabled:
                    available_disabled += 1
            elif status == 'unavailable':
                unavailable += 1
            elif status == 'offline':
                offline += 1
                if disabled:
                    offline_disabled += 1
            elif status == 'unknown':
                unknown += 1
                if disabled:
                    unknown_disabled += 1
        return total, available, available_disabled, unavailable, offline, offline_disabled, unknown, unknown_disabled

    # Virtual Servers
    vs_url = f"https://{address}/mgmt/tm/ltm/virtual/stats"
    vs_stats_json = get_stats(vs_url, token)
    vs_total, vs_avail, vs_avail_dis, vs_unavail, vs_off, vs_off_dis, vs_unk, vs_unk_dis = parse_stats_entries(vs_stats_json)
    # Pools
    pool_url = f"https://{address}/mgmt/tm/ltm/pool/stats"
    pool_stats_json = get_stats(pool_url, token)
    pool_total, pool_avail, pool_avail_dis, pool_unavail, pool_off, pool_off_dis, pool_unk, pool_unk_dis = parse_stats_entries(pool_stats_json)
    # Nodes
    node_url = f"https://{address}/mgmt/tm/ltm/node/stats"
    node_stats_json = get_stats(node_url, token)
    node_total, node_avail, node_avail_dis, node_unavail, node_off, node_off_dis, node_unk, node_unk_dis = parse_stats_entries(node_stats_json)
    # Build summary DataFrame
    summary_df = pd.DataFrame([
        {
            "Object Type": "Virtual Servers",
            "Total": vs_total,
            "Available": f"{vs_avail} ({vs_avail_dis} Disabled)",
            "Unavailable": vs_unavail,
            "Offline": f"{vs_off} ({vs_off_dis} Disabled)",
            "Unknown": f"{vs_unk} ({vs_unk_dis} Disabled)"
        },
        {
            "Object Type": "Pools",
            "Total": pool_total,
            "Available": f"{pool_avail} ({pool_avail_dis} Disabled)",
            "Unavailable": pool_unavail,
            "Offline": f"{pool_off} ({pool_off_dis} Disabled)",
            "Unknown": f"{pool_unk} ({pool_unk_dis} Disabled)"
        },
        {
            "Object Type": "Nodes",
            "Total": node_total,
            "Available": f"{node_avail} ({node_avail_dis} Disabled)",
            "Unavailable": node_unavail,
            "Offline": f"{node_off} ({node_off_dis} Disabled)",
            "Unknown": f"{node_unk} ({node_unk_dis} Disabled)"
        }
    ])
    add_status_summary_to_excel(excel_path, summary_df, device_address=address) 
